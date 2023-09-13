from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
from urllib.parse import urlparse
from time import sleep
import openpyxl
import re


def saveToExcel(products, mallName):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws["A1"] = "Title"
    ws["B1"] = "Price"
    ws["C1"] = "Image"
    ws["D1"] = "Product Link"

    for product in products:
        ws.append(
            [
                product["name"],
                product["price"],
                product["image"],
                product["productLink"],
            ]
        )

    wb.save(f"./output/{mallName}.xlsx")
    print(f"Successfully saved to {mallName}.xlsx")


def scrapeProduct(productLinkList):
    products = []

    for productLink in productLinkList:
        while True:
            try:
                chrome_options = Options()
                chrome_options.add_argument("--no-sandbox")
                driver = webdriver.Chrome(options=chrome_options)
                driver.get(productLink)

                productId = productLink.split("/")[-1]
                print(f"Scraping product: {productId}...")

                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "div._3k440DUKzy"))
                )

                productName = driver.find_element(
                    By.CSS_SELECTOR, "h3._22kNQuEXmb"
                ).text
                productPrice = driver.find_element(
                    By.CSS_SELECTOR, "div._3my-5FC8OB"
                ).text
                productImage = driver.find_element(
                    By.CSS_SELECTOR, "img.bd_2DO68"
                ).get_attribute("src")

                productPrice = re.findall(r"\d{1,3}(?:,\d{3})*(?:\.\d+)?", productPrice)
                prices = [int(prices.replace(",", "")) for prices in productPrice]
                temp = {
                    "name": productName,
                    "price": prices[0],
                    "image": productImage,
                    "productLink": productLink,
                }
                products.append(temp)
                driver.quit()
                break
            except Exception as e:
                print(f"Retrying...{e}\n")
                driver.quit()
                continue

    return products


def scrape(url, end_page, driver):
    driver.get(url)

    wait = WebDriverWait(driver, 30)

    productLinkList = []
    curr_page = 1

    while True:
        if curr_page > end_page:
            break

        for curr in range(1, 11):
            paginationButton = wait.until(
                EC.element_to_be_clickable(
                    (
                        By.CSS_SELECTOR,
                        f"a.UWN4IvaQza:nth-child({curr+1})",
                    )
                )
            )
            paginationButton.click()

            wait.until(
                lambda d: paginationButton.get_attribute("aria-current") == "true"
            )
            print(f"Scraping page {curr_page}...")

            productContainer = wait.until(
                EC.presence_of_element_located(
                    (By.XPATH, "/html/body/div/div/div[3]/div[2]/div[2]/div/ul")
                )
            )

            products = productContainer.find_elements(By.TAG_NAME, "li")

            for product in products:
                productLink = product.find_element(By.TAG_NAME, "a").get_attribute(
                    "href"
                )
                productLinkList.append(productLink)
            url = driver.current_url

            curr_page += 1

            if curr_page > end_page:
                break

        if curr_page <= end_page:
            nextButton = wait.until(
                EC.element_to_be_clickable(
                    (
                        By.XPATH,
                        "/html/body/div/div/div[3]/div[2]/div[2]/div/div[3]/a[12]",
                    )
                )
            )
            nextButton.click()
            print("Getting to next page...")
            sleep(20)

    return productLinkList


def main():
    file = open("./links.txt", "r")
    links = file.readlines()
    file.close()

    for link in links:
        mallLink, ltPg = link.split(" ")
        if ltPg == "all":
            ltPg = 1000

        productLinkList = []

        while True:
            try:
                chrome_options = Options()
                chrome_options.add_argument("--no-sandbox")
                driver = webdriver.Chrome(options=chrome_options)

                productLinks = scrape(mallLink, int(ltPg), driver)

                for productLink in productLinks:
                    productLinkList.append(productLink)

                driver.quit()
                break
            except:
                print(f"\nRetrying...\n")
                driver.quit()
                continue

        mallLName = urlparse(mallLink)
        mallLName = mallLName.path.split("/")[1]
        productsData = scrapeProduct(productLinkList)
        saveToExcel(productsData, mallLName)


if __name__ == "__main__":
    main()
